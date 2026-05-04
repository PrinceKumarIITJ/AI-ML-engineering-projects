"""
The Report Generator (Excel Exporter).
This script takes the final, perfect data and creates a beautiful Excel file.
It automatically creates a separate tab for each city (e.g., one tab for Jaipur, 
one for Delhi), freezes the top row, makes the header blue, and auto-sizes 
the columns so it's instantly ready to be emailed to the sales team.
"""
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List
from models.schema import BusinessSchema
from config import OUTPUT_DIR, EXCEL_COLUMNS

logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self, output_filename: str = "Wedding_Event_Companies_Master.xlsx"):
        self.filepath = OUTPUT_DIR / output_filename

    def _auto_format_sheet(self, worksheet):
        """Applies headers, freezes top row, and attempts auto-sizing."""
        # 1. Freeze top row
        worksheet.freeze_panes = 'A2'
        
        # 2. Style Headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 3. Add Autofilter
        worksheet.auto_filter.ref = worksheet.dimensions

        # 4. Auto-size columns (approximate)
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50) # Cap at 50 width
            worksheet.column_dimensions[column].width = adjusted_width

    def _map_to_excel_row(self, r: BusinessSchema, idx: int) -> dict:
        """Maps a BusinessSchema object to the final Excel presentation format."""
        return {
            "S.No": idx + 1,
            "Business Name": r.business_name or "",
            "Category": r.category or "",
            "City": r.city or "",
            "State": r.state or "",
            "Full Address": r.full_address or "",
            "Pincode": r.pincode or "",
            "Contact Number": r.contact_number or "",
            "Alternate Number": r.alternate_number or "",
            "Email": r.email or "",
            "Website": r.website or "",
            "Instagram": r.instagram_url or "",
            "Facebook": r.facebook_url or "",
            "LinkedIn": r.linkedin_url or "",
            "Google Maps Link": r.google_maps_url or "",
            "Rating": r.rating if r.rating is not None else "",
            "Reviews": r.review_count if r.review_count is not None else "",
            "Price Range": r.price_range or "",
            "Services Offered": ", ".join(r.services_offered) if r.services_offered else "",
            "Years in Business": r.years_in_business if r.years_in_business is not None else "",
            "Source Platforms": ", ".join(r.source_platforms) if r.source_platforms else "",
            "Confidence Score": r.confidence_score or ""
        }

    def export(self, records: List[BusinessSchema]):
        """Creates the city-wise excel file."""
        logger.info(f"Exporting {len(records)} records to {self.filepath}...")
        
        # Sort records by confidence and rating
        def sort_key(r):
            conf_vals = {"verified": 4, "enriched": 3, "raw": 2, "low_confidence": 1}
            return (conf_vals.get(r.confidence_score, 0), r.rating or 0.0)
            
        records.sort(key=sort_key, reverse=True)

        dict_records = [self._map_to_excel_row(r, i) for i, r in enumerate(records)]
        df = pd.DataFrame(dict_records)
        
        # Enforce column order based on config (or fallback to dict keys)
        ordered_cols = [col for col in EXCEL_COLUMNS if col in df.columns]
        for col in [c for c in df.columns if c not in ordered_cols]:
            ordered_cols.append(col)
            
        if not df.empty:
            df = df[ordered_cols]
        else:
            df = pd.DataFrame(columns=ordered_cols)
            
        # Write to separate city tabs
        with pd.ExcelWriter(self.filepath, engine='openpyxl') as writer:
            # Main Summary Tab
            if not df.empty:
                summary_data = {
                    "Total Verified": len(df[df['Confidence Score'] == 'verified']),
                    "Total Enriched": len(df[df['Confidence Score'] == 'enriched']),
                    "Total Raw": len(df[df['Confidence Score'] == 'raw']),
                    "Total Low Confidence": len(df[df['Confidence Score'] == 'low_confidence']),
                    "Total Output": len(df),
                }
            else:
                summary_data = {"Total Output": 0}
                
            summary_df = pd.DataFrame([summary_data]).T
            summary_df.columns = ["Counts"]
            summary_df.to_excel(writer, sheet_name="Summary")

            if not df.empty:
                cities = df['City'].dropna().unique()
                for city in cities:
                    city_df = df[df['City'] == city]
                    safe_city = str(city)[:31].replace('/', '_').replace('\\', '_').replace('?', '').replace('*', '').replace('[', '').replace(']', '')
                    if not safe_city: safe_city = "Unknown"
                    city_df.to_excel(writer, sheet_name=safe_city, index=False)
            else:
                df.to_excel(writer, sheet_name="Data", index=False)

        # Post-process with openpyxl formatting
        try:
            wb = load_workbook(self.filepath)
            for sheet_name in wb.sheetnames:
                self._auto_format_sheet(wb[sheet_name])
            wb.save(self.filepath)
            logger.info("Excel formatting applied successfully.")
        except Exception as e:
            logger.error(f"Error applying Excel styles: {e}")
